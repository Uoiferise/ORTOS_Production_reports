from openpyxl.worksheet import worksheet
from excel_options import MONTH_DICT


def create_production_date(input_sheet: worksheet, row: int, main_file_name: str, month_dict: dict = MONTH_DICT) -> str:

    result = ''
    start_column = (27, 21)[main_file_name == 'input_data/implants/implants_info_new.xlsx']

    if input_sheet.cell(row=row, column=start_column).value is None:
        return result
    else:
        dates = []
        flag = True
        for col in range(start_column + 2, input_sheet.max_column + 1):
            if flag:
                if input_sheet.cell(row=row, column=col).value is None:
                    continue
                else:
                    dates.append([str(input_sheet.cell(row=1, column=col).value)])
                    flag = False
            else:
                if input_sheet.cell(row=row, column=col).value is None:
                    flag = True
                    continue
                else:
                    dates[-1].append(str(input_sheet.cell(row=1, column=col).value))

        for item in dates:
            if result:
                if len(item) == 1:
                    result += f', {item[0].split()[0]}.{month_dict.get(item[0].split()[1])}'
                else:
                    if item[0].split()[-1] == item[-1].split()[-1]:
                        result += f', {item[0].split()[0]}-' \
                                  f'{item[-1].split()[0]}.' \
                                  f'{month_dict.get(item[-1].split()[1])}'
                    else:
                        result += f', {item[0].split()[0]}.{month_dict.get(item[0].split()[1])}-' \
                                  f'{item[-1].split()[0]}.{month_dict.get(item[-1].split()[1])}'
            else:
                if len(item) == 1:
                    result += f'{item[0].split()[0]}.{month_dict.get(item[0].split()[1])}'
                else:
                    if item[0].split()[-1] == item[-1].split()[1]:
                        result += f'{item[0].split()[0]}-' \
                                  f'{item[-1].split()[0]}.' \
                                  f'{month_dict.get(item[-1].split()[1])}'
                    else:
                        result += f'{item[0].split()[0]}.{month_dict.get(item[0].split()[1])}-' \
                                  f'{item[-1].split()[0]}.{month_dict.get(item[-1].split()[1])}'

    return result
