import openpyxl
from excel_options import COLUMN_INDEXES_INFO_DATA, PRODUCTION_PLAN_COLUMN, OTK_COLUMN
from services.basic_report.basic_nomenclature import Nomenclature
from excel_handler.production_plan_handler import create_production_date
from functools import cache


def read_main_file(main_file: str) -> dict:
    main_dict = dict()

    input_book = openpyxl.load_workbook(main_file)
    input_sheet = input_book.active
    for row in range(4, input_sheet.max_row + 1):
        nomenclature_name = input_sheet.cell(row=row, column=5).value
        nomenclature_info = dict()
        for index, col in enumerate(COLUMN_INDEXES_INFO_DATA):
            nomenclature_info[index + 1] = input_sheet.cell(row=row, column=col).value
        nomenclature_info[PRODUCTION_PLAN_COLUMN] = create_production_date(input_sheet=input_sheet, row=row)
        main_dict[nomenclature_name] = Nomenclature(name=nomenclature_name, id_row=row, info=nomenclature_info)

    print(f'{main_file} is loaded')
    return main_dict


def read_unshipped_file(unshipped_file: str) -> dict:
    unshipped_dict = dict()

    if unshipped_file != 'input_data/implants/implants_stock.xlsx':
        unshipped_book = openpyxl.load_workbook(unshipped_file)
        unshipped_sheet = unshipped_book.active

        if unshipped_sheet.max_column == 8:
            for row in range(7, unshipped_sheet.max_row + 1):
                unshipped_dict[str(unshipped_sheet.cell(row=row, column=1).value)] = \
                    int(unshipped_sheet.cell(row=row, column=8).value)
        else:
            print(f'{unshipped_file} have error: max_column != 8:')

    print(f'{unshipped_file} is loaded')
    return unshipped_dict


@cache
def read_otk_file(otk_file: str) -> dict:
    otk_dict = dict()

    input_book = openpyxl.load_workbook(otk_file, data_only=True)
    input_sheet = input_book.active
    for r in range(1, input_sheet.max_row + 1):
        nomenclature_name = input_sheet.cell(row=r, column=1).value
        otk_value = input_sheet.cell(row=r, column=2).value
        otk_dict[nomenclature_name] = otk_value

    print(f'{otk_file} is loaded')
    return otk_dict


def read_data(main_file: str, unshipped_file: str, otk_file: str) -> dict:
    data = read_main_file(main_file=main_file)
    otk_dict = read_otk_file(otk_file=otk_file)
    unshipped_dict = read_unshipped_file(unshipped_file=unshipped_file)

    for nomenclature_name in data.keys():
        nomenclature_name_info = data[nomenclature_name].get_info()
        nomenclature_name_info[OTK_COLUMN] = otk_dict.get(nomenclature_name, None)
        nomenclature_name_info[OTK_COLUMN + 1] = None
        nomenclature_name_info[OTK_COLUMN + 2] = unshipped_dict.get(nomenclature_name, None)

    return data
