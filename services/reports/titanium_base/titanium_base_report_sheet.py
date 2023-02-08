from services.basic_report.basic_report_sheet import BasicReportSheet
from openpyxl.workbook import Workbook


class TBReportSheet(BasicReportSheet):
    __slots__ = ()

    @staticmethod
    def validation_data(data: dict) -> tuple:
        return data, data

    def __init__(self, wb: Workbook, name: str, data):
        super().__init__(wb, name, data)
        self._data_bridge, self._data_single = self.validation_data(data=self._data)

    def transport_date(self, data: dict) -> None:
        row = self._start_row
        for nomenclature in self._data.values():
            nomenclature_info = nomenclature.get_info()
            for col in range(1, len(nomenclature_info) + 1):
                cell = self._sheet.cell(row=row, column=col)
                info = nomenclature_info[col]
                if 9 <= col <= 18 and info is None:
                    info = 0
                cell.value = info
                self.cell_style(cell)
            row += 1

    def create_sheet(self) -> None:
        self.create_sheet_header()
        self.transport_date(self._data_single)
        self.fill_small_stock()
        self.separation_nomenclatures()
        self.create_sheet_resul()
