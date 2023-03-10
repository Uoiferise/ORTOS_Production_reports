import openpyxl
from services.basic_report.basic_report import BasicReport
from services.basic_report.basic_report_sheet import BasicReportSheet
from settings import REPORTS_NAME_DICT
from openpyxl.workbook import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


class ReportTitaniumBase(BasicReport):
    __slots__ = ()

    __TB_YELLOW_NOMENCLATURES = (
        '38779',
        '38780',
        '38781',
        '38711',
        '38732',
        '38735',
        '38737',
        '38755',
        '38625',
        '38626',
        '38621',
        '38622',
        '38624',
        '38606',
        '38608',
        '38617',
        '38618'
    )

    __TB_ACTUAL_BOOK_PATH = 'services/reports/titanium_base/titanium_base_actual.xlsx'

    def validation_data(self, data: dict) -> dict:
        # Aggregation of information of some archival nomenclatures
        tb_actual_book = openpyxl.load_workbook(self.__TB_ACTUAL_BOOK_PATH)
        tb_actual_sheet = tb_actual_book.active
        for r in range(2, tb_actual_sheet.max_row + 1):
            actual_nom_name = tb_actual_sheet.cell(row=r, column=1).value
            archival_nom_name = tb_actual_sheet.cell(row=r, column=2).value
            if actual_nom_name in data.keys() and archival_nom_name in data.keys():
                add_info = data[archival_nom_name].get_info()
                data[actual_nom_name].aggregate_info(add_info)

        # Selection of non-archival items in data and delete some nomenclatures
        data_copy = data.copy()
        for key, value in data_copy.items():
            if value.get_info()[6] == 'Да' or \
                    value.get_info()[7] == 'Да' or \
                    'струк' in value.get_info()[5].lower() or \
                    '2к' in value.get_info()[5] or \
                    'кат2' in value.get_info()[5].split()[0]:
                del data[key]

        return data

    @staticmethod
    def grouping_of_nomenclatures(data: dict, conditions: tuple) -> dict:
        aggr_nom = {}
        for nom_name, nomenclature in data.items():
            if any(item in nomenclature.get_info()[8] for item in conditions):
                aggr_nom[nom_name] = nomenclature

        for nom_name in aggr_nom.keys():
            del data[nom_name]

        for nom_name, nomenclature in data.items():
            for item in aggr_nom.values():
                if item.get_info()[8][:-2] == nomenclature.get_info()[8] and \
                        item.vendor_code == nomenclature.vendor_code:
                    nomenclature.aggregate_info(item.get_info())

        return data

    def create_report(self) -> None:
        # divide the information into 7 sheets
        data_1 = dict()
        data_2 = dict()
        data_3 = dict()
        data_4 = dict()
        data_5 = dict()
        data_6 = dict()
        data_7 = dict()

        for key, value in self._data.items():
            if 'Patch' == value.get_info()[2]:
                data_1[key] = value
            elif 'Flat' == value.get_info()[2]:
                data_2[key] = value
            elif 'Half' == value.get_info()[2]:
                data_3[key] = value
            elif 'Bell GEO' == value.get_info()[2]:
                data_4[key] = value
            elif 'Step GEO' == value.get_info()[2]:
                data_5[key] = value
            elif 'Step ARUM' == value.get_info()[2]:
                data_6[key] = value
            else:
                data_7[key] = value

        sheets_dict = {
            'Остальное': (BasicReportSheet, data_7),
            'Arum': (BasicReportSheet, data_6),
            'GEO Step': (BasicReportSheet, data_5),
            'GEO Bell': (TBReportSheet, data_4),
            'Half (ИМ Абатменты.ру)': (BasicReportSheet, self.grouping_of_nomenclatures(data_3, ('P', 'Н'))),
            'Flat с насечками (ИМ Ортос)': (TBReportSheet, self.grouping_of_nomenclatures(data_2, ('P', ))),
            'Patch': (TBReportSheet, self.grouping_of_nomenclatures(data_1, ('P', 'Н'))),
        }

        for name, value in sheets_dict.items():
            current_sheet = value[0](wb=self._workbook, name=name, data=value[1])
            current_sheet.create_sheet()
            if name in ['GEO Step', 'Arum']:
                sheet = current_sheet.get_sheet()
                for row in range(7, sheet.max_row):
                    cell = sheet.cell(row=row, column=5)
                    if cell.value is not None and cell.value.split()[0][:5] in self.__TB_YELLOW_NOMENCLATURES:
                        cell.fill = PatternFill("solid", fgColor="FFFF99")

        self._workbook.save(filename=REPORTS_NAME_DICT[self.report_name]['report_name'])


class TBReportSheet(BasicReportSheet):
    __slots__ = ('_data_bridge', '_data_single')

    @staticmethod
    def divide_data(data: dict) -> tuple:
        data_bridge, data_single = dict(), dict()
        for key, value in data.items():
            if 'ТО bridge' == value.get_info()[1]:
                data_bridge[key] = value
            else:
                data_single[key] = value
        return data_bridge, data_single

    def __init__(self, wb: Workbook, name: str, data):
        super().__init__(wb, name, data)
        self._data_bridge, self._data_single = self.divide_data(data=self._data)

    def add_title(self, title: str, row: int) -> None:
        cell = self._sheet.cell(row=row, column=5)
        cell.value = title
        for col in range(1, self._sheet.max_column + 1):
            cell = self._sheet.cell(row=row, column=col)
            self.cell_style(cell)
            if col == 5:
                cell.font = Font(name='Arial', bold=True, size=10)
                cell.alignment = Alignment(horizontal='center', vertical='center')

    def create_sheet(self) -> None:
        self.create_sheet_header()

        self.add_title(title='Мостовидные', row=self._sheet.max_row)
        start_row = self._sheet.max_row + 1
        self.transport_date(self._data_bridge, start_row=start_row)
        self.separation_nomenclatures(start_row=start_row)
        for col in range(1, self._sheet.max_column + 1):
            cell = self._sheet.cell(row=self._sheet.max_row, column=col)
            self.cell_style(cell)

        self.add_title(title='Одиночные', row=self._sheet.max_row+1)
        start_row = self._sheet.max_row + 1
        self.transport_date(self._data_single, start_row=start_row)
        self.separation_nomenclatures(start_row=start_row)

        self.fill_small_stock()
        self.create_sheet_resul()
