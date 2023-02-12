from services.basic_report.basic_report import BasicReport
from services.basic_report.basic_report_sheet import BasicReportSheet
from settings import REPORTS_NAME_DICT
import openpyxl
from copy import copy


class ReportBlanks(BasicReport):
    __slots__ = ()

    def create_report(self):
        # divide the information into 6 sheets
        data_1 = dict()
        data_2 = dict()
        data_3 = dict()
        data_4 = dict()
        data_5 = dict()

        for key, value in self._data.items():
            if 'Для холдеров ADM / MEDENTiKA' == value.get_info()[2]:
                data_4[key] = value
            elif 'Для холдера ZIRKONZAHN' == value.get_info()[2]:
                data_3[key] = value
            elif 'Для холдера ARUM' == value.get_info()[2] and \
                    'КХ' in value.get_info()[5].split()[0]:
                data_1[key] = value
            elif 'Для холдера ARUM' == value.get_info()[2]:
                data_2[key] = value
            elif 'LM Short' not in value.get_info()[5] and 'LM Long' not in value.get_info()[5]:
                data_5[key] = value

        sheets_dict = {
            'Остальное': (BasicReportSheet, data_5),
            'DM_Medentika': (BasicReportSheet, data_4),
            'Zirkonzahn': (BasicReportSheet, data_3),
            'ARUM_Ti': (BasicReportSheet, data_2),
            'ARUM_CoCr': (BasicReportSheet, data_1),
            'LM1': (ReportBlanksSheet, self._data),
        }

        for name, value in sheets_dict.items():
            current_sheet = value[0](wb=self._workbook, name=name, data=value[1])
            current_sheet.create_sheet()

        self._workbook.save(filename=REPORTS_NAME_DICT[self.report_name]['report_name'])


class ReportBlanksSheet(BasicReportSheet):
    __slots__ = ()

    __COPY_SHEET_PATH_DICT = {
        'LM1': 'services/reports/blanks/blanks_LM1.xlsx',
    }

    def copy_sheet(self, wb_path: str) -> None:
        wb_from = openpyxl.load_workbook(wb_path)
        ws_from = wb_from.active

        for i in range(self._sheet.max_row, ws_from.max_row + 1):
            for j in range(1, ws_from.max_column + 1):
                # reading cell value from source excel file
                cell_from = ws_from.cell(row=i, column=j)

                # writing the read value to destination excel file
                self._sheet.cell(row=i, column=j).value = cell_from.value
                new_cell = self._sheet.cell(row=i, column=j)

                if cell_from.has_style:
                    new_cell.font = copy(cell_from.font)
                    new_cell.border = copy(cell_from.border)
                    new_cell.fill = copy(cell_from.fill)
                    new_cell.number_format = copy(cell_from.number_format)
                    new_cell.protection = copy(cell_from.protection)
                    new_cell.alignment = copy(cell_from.alignment)

            self._sheet.row_dimensions[i].height = 15

    def transport_date(self, data: dict, start_row: int = None) -> None:
        for row in range(7, self._sheet.max_row + 1):
            nomenclature_name = self._sheet.cell(row=row, column=5).value
            if data.get(nomenclature_name, None) is not None:
                for col in range(1, self._sheet.max_column + 1):
                    cell = self._sheet.cell(row=row, column=col)
                    info = data[nomenclature_name].get_info()[col]
                    cell.value = info
        self._sheet.column_dimensions.group('A', 'D', hidden=True)

    def create_sheet(self) -> None:
        self.create_sheet_header()
        self.copy_sheet(wb_path=self.__COPY_SHEET_PATH_DICT[self.name])
        self.transport_date(data=self._data)
        self.fill_small_stock()
        self.create_sheet_resul()
