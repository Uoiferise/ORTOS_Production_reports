from services.basic_report.basic_report import BasicReport
from services.basic_report.basic_report_sheet import BasicReportSheet
from settings import REPORTS_NAME_DICT
import openpyxl
from openpyxl.styles import PatternFill
from copy import copy
from functools import cache


class ReportImplants(BasicReport):
    __slots__ = ()

    def validation_data(self, data: dict) -> dict:
        return data

    def create_report(self):

        sheets_dict = {
            'Implantium': (ReportImplantsSheet, self._data),
            'Nobel Active': (ReportImplantsSheet, self._data),
            'Osstem': (OsstemSheet, self._data),
        }

        for name, value in sheets_dict.items():
            current_sheet = value[0](wb=self._workbook, name=name, data=value[1])
            current_sheet.create_sheet()

        self._workbook.save(filename=REPORTS_NAME_DICT[self.report_name]['report_name'])


@cache
def read_otk_file(otk_file: str) -> dict:
    otk_dict = dict()

    input_book = openpyxl.load_workbook(otk_file, data_only=True)
    input_sheet = input_book.active
    for r in range(1, input_sheet.max_row + 1):
        nomenclature_name = input_sheet.cell(row=r, column=1).value
        otk_value = input_sheet.cell(row=r, column=2).value
        otk_dict[nomenclature_name] = otk_value

    return otk_dict


class ReportImplantsSheet(BasicReportSheet):
    __slots__ = ()

    _HEADERS_DICT = {
            1: 'Тип',
            2: 'Линейка',
            3: 'Система',
            4: 'Разм',
            5: 'Номенклатура',
            6: 'Арх ном',
            7: 'Арх кат',
            8: 'Карт кат',
            9: 'Остаток',
            10: 'Расход общий',
            19: 'ПЛАН',
            20: 'Произведено / неоприходовано',
            21: 'Непроизведено / в плане',
            22: 'Неотгружено по опт. заявкам',
            23: 'Необходимый остаток на 4 месяца',
            24: 'Оригинал',
            25: 'КД',
    }

    _OTK_DATA = read_otk_file(otk_file='input_data/otk.xlsx')

    __COPY_SHEET_PATH_DICT = {
        'Implantium': 'services/reports/implants/Implantium.xlsx',
        'Nobel Active': 'services/reports/implants/Nobel Active.xlsx',
        'Osstem': 'services/reports/implants/Osstem.xlsx',
    }

    def copy_sheet(self, wb_path: str) -> None:
        wb_from = openpyxl.load_workbook(wb_path)
        ws_from = wb_from.active

        for i in range(self._sheet.max_row, ws_from.max_row + 1):
            for j in range(1, ws_from.max_column + 1):
                # reading cell value from source excel file
                cell_from = ws_from.cell(row=i, column=j)

                # writing the read value to destination excel file
                self._sheet.cell(row=i, column=j).value = cell_from.value
                new_cell = self._sheet.cell(row=i, column=j)

                if cell_from.has_style:
                    new_cell.font = copy(cell_from.font)
                    new_cell.border = copy(cell_from.border)
                    new_cell.fill = copy(cell_from.fill)
                    new_cell.number_format = copy(cell_from.number_format)
                    new_cell.protection = copy(cell_from.protection)
                    new_cell.alignment = copy(cell_from.alignment)

            self._sheet.row_dimensions[i].height = 15

    def transport_date(self, data: dict, start_row: int = None) -> None:
        for row in range(7, self._sheet.max_row + 1):
            nomenclature_name = self._sheet.cell(row=row, column=5).value
            if data.get(nomenclature_name, None) is not None:
                for col in range(9, 22):
                    cell = self._sheet.cell(row=row, column=col)
                    info = data[nomenclature_name].get_info().get(col, None)
                    if 9 <= col <= 18 and info is None:
                        info = 0
                    cell.value = info
            elif nomenclature_name is not None and nomenclature_name != 'Итого':
                for col in range(9, 22):
                    cell = self._sheet.cell(row=row, column=col)
                    if 9 <= col <= 18:
                        info = 0
                    elif col == 20:
                        info = self._OTK_DATA.get(nomenclature_name, None)
                    else:
                        info = None
                    cell.value = info
        self._sheet.column_dimensions.group('A', 'D', hidden=True)
        self._sheet.column_dimensions.group('K', 'R', hidden=True)

    def fill_columns(self) -> None:
        for col in (24, 25):
            for row in range(self._start_row, self._sheet.max_row):
                cell = self._sheet.cell(row=row, column=col)
                value = cell.value
                if value is None:
                    continue
                elif value in ('есть', 'не требуется'):
                    cell.fill = PatternFill("solid", fgColor="C6EFCE")
                elif value == 'в разработке':
                    cell.fill = PatternFill("solid", fgColor="FFEB9C")
                else:
                    cell.fill = PatternFill("solid", fgColor="FFC7CE")

    def create_sheet(self) -> None:
        self.create_sheet_header()
        self.copy_sheet(wb_path=self.__COPY_SHEET_PATH_DICT[self.name])
        self.transport_date(data=self._data)
        self.fill_columns()


class OsstemSheet(ReportImplantsSheet):
    __slots__ = ()

    _HEADERS_DICT = {
            1: 'Тип',
            2: 'Линейка',
            3: 'Система',
            4: 'Разм',
            5: 'Номенклатура',
            6: 'Арх ном',
            7: 'Арх кат',
            8: 'Карт кат',
            9: 'Остаток',
            10: 'Расход общий',
            19: 'ПЛАН',
            20: 'Произведено / неоприходовано',
            21: 'Непроизведено / в плане',
            22: 'Неотгружено по опт. заявкам',
            23: 'Необходимый остаток на 4 месяца',
    }
