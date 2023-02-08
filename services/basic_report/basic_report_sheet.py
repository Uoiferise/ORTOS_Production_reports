from services.abstractions.abstract_report_sheet import AbstractReportSheet
from openpyxl import Workbook
from basic_settings import HEADERS_DICT
from settings import DATE_START, DATE_STOP
from basic_functions.create_sheet_header import create_sheet_header
from basic_functions.fill_small_stock import fill_small_stock
from basic_functions.create_sheet_result import create_sheet_result
from basic_functions.separation_nomenclatures import separation_nomenclatures
from openpyxl.styles import Alignment


class BasicReportSheet(AbstractReportSheet):
    """The class responsible for creating the report sheet"""

    __slots__ = ('_sheet', '_data', '_start_row')

    _HEADERS_DICT = HEADERS_DICT
    _DATE_START = DATE_START
    _DATE_STOP = DATE_STOP

    def __init__(self, wb: Workbook, name: str, data):
        self._sheet = wb.create_sheet(title=name, index=0)
        self._data = data
        self._start_row = self._sheet.max_row

    def create_sheet_header(self) -> None:
        create_sheet_header(sheet=self._sheet,
                            date_start=self._DATE_START,
                            date_stop=self._DATE_STOP,
                            header_dict=self._HEADERS_DICT)
        self._start_row = self._sheet.max_row

    def fill_small_stock(self) -> None:
        fill_small_stock(sheet=self._sheet,
                         start_row=self._start_row)

    def create_sheet_resul(self) -> None:
        create_sheet_result(sheet=self._sheet,
                            start_row=self._start_row,
                            end_row=self._sheet.max_row)

    def separation_nomenclatures(self) -> None:
        separation_nomenclatures(sheet=self._sheet,
                                 start_row=self._start_row)

    @staticmethod
    def cell_style(cell) -> None:
        col = cell.column
        if col <= 18 or col == 22:
            cell.style = 'info'
        elif col == 19:
            cell.style = 'date'
        elif col == 20 or col == 21:
            cell.style = 'white'
        if col >= 9:
            cell.alignment = Alignment(horizontal='center', vertical='center')

    def transport_date(self, data: dict, start_row: int = None) -> None:
        row = (start_row, self._start_row)[start_row is None]
        for nomenclature in data.values():
            nomenclature_info = nomenclature.get_info()
            for col in range(1, len(nomenclature_info) + 1):
                cell = self._sheet.cell(row=row, column=col)
                info = nomenclature_info[col]
                if 9 <= col <= 18 and info is None:
                    info = 0
                cell.value = info
                self.cell_style(cell)
            row += 1

    def create_sheet(self) -> None:
        self.create_sheet_header()
        self.transport_date(data=self._data)
        self.fill_small_stock()
        self.separation_nomenclatures()
        self.create_sheet_resul()
