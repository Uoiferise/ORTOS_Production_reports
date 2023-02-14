from openpyxl.styles import Alignment, PatternFill


def fill_small_stock(sheet, start_row: int) -> None:
    for row in range(start_row, sheet.max_row + 1):
        if sheet.cell(row=row, column=9).value is None:
            continue
        elif sheet.cell(row=row, column=10).value >= sheet.cell(row=row, column=9).value:
            sheet.cell(row=row, column=5).fill = PatternFill("solid", fgColor="FFCCCC")
            for c in [9, 10]:
                sheet.cell(row=row, column=c).fill = PatternFill("solid", fgColor="FFCCCC")
                sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
        elif sheet.cell(row=row, column=13).value != 0 and isinstance(sheet.cell(row=row, column=11).value, int):
            if (sheet.cell(row=row, column=11).value / sheet.cell(row=row, column=13).value) < 1:
                sheet.cell(row=row, column=5).fill = PatternFill("solid", fgColor="CCECFF")
                for c in [11, 13]:
                    sheet.cell(row=row, column=c).fill = PatternFill("solid", fgColor="CCECFF")
                    sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
            elif 1 <= (sheet.cell(row=row, column=11).value / sheet.cell(row=row, column=13).value) < 2.5:
                sheet.cell(row=row, column=5).fill = PatternFill("solid", fgColor="CCFFCC")
                for c in [11, 13]:
                    sheet.cell(row=row, column=c).fill = PatternFill("solid", fgColor="CCFFCC")
                    sheet.cell(row=row, column=c).alignment = Alignment(horizontal='center', vertical='center')
