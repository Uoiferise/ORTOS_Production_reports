from openpyxl.styles import Font
from openpyxl.utils import get_column_letter


def create_sheet_header(sheet, date_start: str, date_stop: str, header_dict: dict) -> None:

    sheet.cell(row=1, column=5).value = f'Конец периода: {date_start} 23:59:59'
    sheet.cell(row=1, column=5).font = Font(name='Arial', bold=False, size=8)

    sheet.cell(row=2, column=5).value = f'Начало периода: {date_stop} 00:00:00'
    sheet.cell(row=2, column=5).font = Font(name='Arial', bold=False, size=8)

    for key in header_dict.keys():
        if 9 <= key <= 10:
            sheet.merge_cells(start_row=5, start_column=int(key), end_row=6, end_column=int(key))
            sheet.cell(row=5, column=int(key)).value = header_dict[key]
        else:
            sheet.merge_cells(start_row=4, start_column=int(key), end_row=6, end_column=int(key))
            sheet.cell(row=4, column=int(key)).value = header_dict[key]

    sheet.merge_cells(start_row=4, start_column=9, end_row=4, end_column=10)
    sheet.cell(row=4, column=9).value = 'Итого'
    sheet.merge_cells(start_row=4, start_column=11, end_row=4, end_column=14)
    sheet.cell(row=4, column=11).value = 'ОСНОВНЫЕ СКЛАДЫ'
    sheet.merge_cells(start_row=4, start_column=15, end_row=4, end_column=18)
    sheet.cell(row=4, column=15).value = 'ПРОЧИЕ СКЛАДЫ'

    for c in [11, 15]:
        sheet.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c + 1)
        sheet.cell(row=5, column=c).value = 'ОСТ'

    for c in [13, 17]:
        sheet.merge_cells(start_row=5, start_column=c, end_row=5, end_column=c + 1)
        sheet.cell(row=5, column=c).value = 'РАСХ'

    for c in range(11, 19):
        if c % 2 != 0:
            sheet.cell(row=6, column=c).value = 'ИЗД'
        else:
            sheet.cell(row=6, column=c).value = 'К/Т'

    for row in range(4, 7):
        for c in range(1, sheet.max_column + 1):
            sheet.cell(row=row, column=c).style = 'header'

    for c in range(1, sheet.max_column + 1):
        if c <= 4:
            sheet.column_dimensions[get_column_letter(c)].width = 9
        elif c == 5:
            sheet.column_dimensions[get_column_letter(c)].width = 90
        elif 6 <= c <= 8:
            sheet.column_dimensions[get_column_letter(c)].width = 7.5
        elif 9 <= c <= 18:
            sheet.column_dimensions[get_column_letter(c)].width = 8.25
        else:
            sheet.column_dimensions[get_column_letter(c)].width = 20

    sheet.column_dimensions.group('F', 'H', hidden=True)
    sheet.freeze_panes = sheet.cell(row=7, column=6)
