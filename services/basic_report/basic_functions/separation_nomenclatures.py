from openpyxl.styles import Side, PatternFill, Border


def separation_nomenclatures(sheet, start_row: int, exception: bool = False) -> None:
    # Будем разделять номенклатуры по их линейке, а также объединять пары с одинаковыми размерами в жирные границы
    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")

    row_indexes_for_new_line = []
    for row in range(start_row, sheet.max_row + 1):
        if sheet.cell(row=row, column=3).value == sheet.cell(row=(row + 1), column=3).value:
            continue
        else:
            row_indexes_for_new_line.append(row + 1)

    for i in enumerate(row_indexes_for_new_line[:-1:]):
        sheet.insert_rows(idx=(i[1] + i[0]))
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=(i[1] + i[0]), column=col)
            if col != 20 and col != 21:
                cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
            else:
                cell.fill = PatternFill("solid", fgColor="FFFFFF")
    if exception:
        sheet.column_dimensions.group('A', 'D', hidden=True)
    else:
        border_highlighting_list = []
        for i in enumerate(row_indexes_for_new_line):
            flag = True
            for row in range(start_row, i[1] + i[0] - 1):
                if flag:
                    if sheet.cell(row=row, column=4).value == sheet.cell(row=(row + 1), column=4).value:
                        continue
                    else:
                        flag = False
                        stop_row = row
                        border_highlighting_list.append([start_row, stop_row])
                        start_row = row + 1
                else:
                    if sheet.cell(row=row, column=4).value == sheet.cell(row=(row + 1), column=4).value and\
                            sheet.cell(row=(row + 2), column=4).value is not None:
                        continue
                    elif sheet.cell(row=row, column=4).value == sheet.cell(row=(row + 1), column=4).value and\
                            sheet.cell(row=(row + 2), column=4).value is None:
                        flag = True
                        stop_row = row + 1
                        border_highlighting_list.append([start_row, stop_row])
                        start_row = row + 1
                    else:
                        stop_row = row
                        border_highlighting_list.append([start_row, stop_row])
                        start_row = row + 1
            start_row = i[1] + i[0] + 1

        # Закрашиваем границы отобранных ячеек
        for item in border_highlighting_list:
            if (item[-1] - item[0]) >= 1:
                for row in range(item[0], item[1] + 1):
                    for col in range(5, 11):
                        cell = sheet.cell(row=row, column=col)
                        if row == item[0]:
                            if col == 5:
                                cell.border = Border(top=medium, left=medium, right=thin, bottom=thin)
                            elif col == 10:
                                cell.border = Border(top=medium, left=thin, right=medium, bottom=thin)
                            else:
                                cell.border = Border(top=medium, left=thin, right=thin, bottom=thin)
                        elif row == item[1]:
                            if col == 5:
                                cell.border = Border(top=thin, left=medium, right=thin, bottom=medium)
                            elif col == 10:
                                cell.border = Border(top=thin, left=thin, right=medium, bottom=medium)
                            else:
                                cell.border = Border(top=thin, left=thin, right=thin, bottom=medium)
                        else:
                            if col == 5:
                                cell.border = Border(top=thin, left=medium, right=thin, bottom=thin)
                            elif col == 10:
                                cell.border = Border(top=thin, left=thin, right=medium, bottom=thin)

        sheet.column_dimensions.group('A', 'D', hidden=True)
