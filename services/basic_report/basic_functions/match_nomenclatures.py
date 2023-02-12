import openpyxl
from openpyxl.styles import Side, Border, PatternFill, Font


def match_nomenclatures(match_file: str, sheet):
    matching_book = openpyxl.load_workbook(match_file, data_only=True)
    matching_sheet = matching_book.active
    matching_dict = {}
    for row in range(2, matching_sheet.max_row + 1):
        key = str(matching_sheet.cell(row=row, column=1).value)
        value = (
            matching_sheet.cell(row=row, column=3).value,
            matching_sheet.cell(row=row, column=5).value
        )
        matching_dict[key] = value

    thin = Side(border_style="thin", color="000000")
    medium = Side(border_style="medium", color="000000")

    row = 7
    while sheet.cell(row=row, column=5).value != "Итого":
        if sheet.cell(row=row, column=5).value is None:
            row += 1
            continue

        key = sheet.cell(row=row, column=5).value.split()[0]
        if matching_dict.get(key, False):
            if sheet.cell(row=row + 1, column=5).value == matching_dict[key][0]:
                if sheet.cell(row=row + 2, column=5).value == matching_dict[key][1]:
                    for i in range(3):
                        for col in range(5, 11):
                            cell = sheet.cell(row=row + i, column=col)
                            if i == 0:
                                if col == 5:
                                    cell.border = Border(top=medium, left=medium, right=thin, bottom=thin)
                                elif col == 10:
                                    cell.border = Border(top=medium, left=thin, right=medium, bottom=thin)
                                else:
                                    cell.border = Border(top=medium, left=thin, right=thin, bottom=thin)
                            elif i == 2:
                                if col == 5:
                                    cell.border = Border(top=thin, left=medium, right=thin, bottom=medium)
                                elif col == 10:
                                    cell.border = Border(top=thin, left=thin, right=medium, bottom=medium)
                                else:
                                    cell.border = Border(top=thin, left=thin, right=thin, bottom=medium)
                            else:
                                if col == 5:
                                    cell.border = Border(top=thin, left=medium, right=thin, bottom=thin)
                                elif col == 10:
                                    cell.border = Border(top=thin, left=thin, right=medium, bottom=thin)
                    row += 3
                else:
                    sheet.insert_rows(idx=row + 2)
                    sheet.cell(row=row + 2, column=5).value = matching_dict[key][1]
                    sheet.cell(row=row + 2, column=5).font = Font(name='Arial', bold=False, size=8)

                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row + 2, column=col)
                        if col != 20 and col != 21:
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        else:
                            cell.fill = PatternFill("solid", fgColor="FFFFFF")

                    for i in range(3):
                        for col in range(5, 11):
                            cell = sheet.cell(row=row + i, column=col)
                            if i == 0:
                                if col == 5:
                                    cell.border = Border(top=medium, left=medium, right=thin, bottom=thin)
                                elif col == 10:
                                    cell.border = Border(top=medium, left=thin, right=medium, bottom=thin)
                                else:
                                    cell.border = Border(top=medium, left=thin, right=thin, bottom=thin)
                            elif i == 2:
                                if col == 5:
                                    cell.border = Border(top=thin, left=medium, right=thin, bottom=medium)
                                elif col == 10:
                                    cell.border = Border(top=thin, left=thin, right=medium, bottom=medium)
                                else:
                                    cell.border = Border(top=thin, left=thin, right=thin, bottom=medium)
                            else:
                                if col == 5:
                                    cell.border = Border(top=thin, left=medium, right=thin, bottom=thin)
                                elif col == 10:
                                    cell.border = Border(top=thin, left=thin, right=medium, bottom=thin)
                    row += 3
            else:
                sheet.insert_rows(idx=row + 1)
                sheet.cell(row=row + 1, column=5).value = matching_dict[key][0]
                sheet.cell(row=row + 1, column=5).font = Font(name='Arial', bold=False, size=8)

                for col in range(1, sheet.max_column + 1):
                    cell = sheet.cell(row=row + 1, column=col)
                    if col != 20 and col != 21:
                        cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                    else:
                        cell.fill = PatternFill("solid", fgColor="FFFFFF")

                if sheet.cell(row=row + 2, column=5).value == matching_dict[key][1]:
                    for i in range(3):
                        for col in range(5, 11):
                            cell = sheet.cell(row=row + i, column=col)
                            if i == 0:
                                if col == 5:
                                    cell.border = Border(top=medium, left=medium, right=thin, bottom=thin)
                                elif col == 10:
                                    cell.border = Border(top=medium, left=thin, right=medium, bottom=thin)
                                else:
                                    cell.border = Border(top=medium, left=thin, right=thin, bottom=thin)
                            elif i == 2:
                                if col == 5:
                                    cell.border = Border(top=thin, left=medium, right=thin, bottom=medium)
                                elif col == 10:
                                    cell.border = Border(top=thin, left=thin, right=medium, bottom=medium)
                                else:
                                    cell.border = Border(top=thin, left=thin, right=thin, bottom=medium)
                            else:
                                if col == 5:
                                    cell.border = Border(top=thin, left=medium, right=thin, bottom=thin)
                                elif col == 10:
                                    cell.border = Border(top=thin, left=thin, right=medium, bottom=thin)
                    row += 3
                else:
                    sheet.insert_rows(idx=row + 2)
                    sheet.cell(row=row + 2, column=5).value = matching_dict[key][1]
                    sheet.cell(row=row + 2, column=5).font = Font(name='Arial', bold=False, size=8)

                    for col in range(1, sheet.max_column + 1):
                        cell = sheet.cell(row=row + 2, column=col)
                        if col != 20 and col != 21:
                            cell.border = Border(top=thin, left=thin, right=thin, bottom=thin)
                        else:
                            cell.fill = PatternFill("solid", fgColor="FFFFFF")

                    for i in range(3):
                        for col in range(5, 11):
                            cell = sheet.cell(row=row + i, column=col)
                            if i == 0:
                                if col == 5:
                                    cell.border = Border(top=medium, left=medium, right=thin, bottom=thin)
                                elif col == 10:
                                    cell.border = Border(top=medium, left=thin, right=medium, bottom=thin)
                                else:
                                    cell.border = Border(top=medium, left=thin, right=thin, bottom=thin)
                            elif i == 2:
                                if col == 5:
                                    cell.border = Border(top=thin, left=medium, right=thin, bottom=medium)
                                elif col == 10:
                                    cell.border = Border(top=thin, left=thin, right=medium, bottom=medium)
                                else:
                                    cell.border = Border(top=thin, left=thin, right=thin, bottom=medium)
                            else:
                                if col == 5:
                                    cell.border = Border(top=thin, left=medium, right=thin, bottom=thin)
                                elif col == 10:
                                    cell.border = Border(top=thin, left=thin, right=medium, bottom=thin)
                    row += 3
        else:
            row += 1
